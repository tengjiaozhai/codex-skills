"""Excel and JSON exporters for full EDIF extraction tables."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .edif_full_extract import FullEdifTables, SHEET_ORDER


FULL_EXPORT_COLUMNS: dict[str, list[str]] = {
    "Summary": [
        "file_path",
        "edif_name",
        "library_count",
        "page_count",
        "instance_count",
        "net_count",
        "property_count",
        "coordinate_count",
        "raw_node_count",
        "raw_property_count",
        "raw_origin_count",
        "raw_pt_count",
        "raw_orientation_count",
        "raw_display_count",
        "raw_library_count",
        "raw_cell_count",
        "raw_interface_count",
        "raw_port_count",
        "raw_portInstance_count",
        "raw_array_count",
        "raw_page_count",
        "raw_instance_count",
        "raw_net_count",
    ],
    "Libraries": ["library_name", "properties_json", "source_path"],
    "Cells": [
        "library_name",
        "cell_name",
        "cell_display_name",
        "cellType",
        "properties_json",
    ],
    "Ports": [
        "library_name",
        "cell_name",
        "view_name",
        "port_name",
        "designator",
        "Type",
        "PackagePortNumbers",
        "properties_json",
    ],
    "Pages": [
        "schematic",
        "page_name",
        "page_token",
        "properties_json",
        "source_path",
    ],
    "Instances": [
        "instance_id",
        "refdes",
        "schematic",
        "page",
        "library_ref",
        "cell_ref",
        "view_ref",
        "origin_x",
        "origin_y",
        "orientation",
        "properties_json",
    ],
    "InstanceProperties": [
        "instance_id",
        "refdes",
        "property_key",
        "display_name",
        "value_raw",
        "value_text",
        "display_origin_x",
        "display_origin_y",
        "visible",
        "source_path",
    ],
    "PinInstances": [
        "instance_id",
        "refdes",
        "symbol_pin",
        "pin_number",
        "pin_name",
        "net_name",
        "schematic",
        "page",
        "origin_x",
        "origin_y",
    ],
    "Nets": [
        "net_name",
        "schematic",
        "page",
        "connection_count",
        "Pins",
        "source_path",
    ],
    "NetConnections": [
        "net_name",
        "refdes",
        "instance_id",
        "symbol_pin",
        "pin_number",
        "pin_name",
        "resolved",
    ],
    "Displays": [
        "owner",
        "display_type",
        "origin_x",
        "origin_y",
        "justify",
        "visible",
        "raw_json",
    ],
    "Geometry": [
        "owner",
        "geometry_type",
        "point_list",
        "origin",
        "orientation",
        "source_path",
    ],
    "Arrays": [
        "array_name",
        "length",
        "owner_library",
        "owner_cell",
        "owner_interface",
        "source_path",
    ],
    "HierarchyRefs": [
        "instance_id",
        "view_ref",
        "cell_ref",
        "library_ref",
        "source_path",
    ],
    "RawNodes": [
        "node_path",
        "parent_path",
        "depth",
        "head",
        "value",
        "line",
        "column",
        "child_index",
    ],
}


def _cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def _sheet_title(base: str, index: int) -> str:
    if index == 1:
        return base[:31]
    suffix = f"_{index}"
    return f"{base[:31 - len(suffix)]}{suffix}"


def _append_table(ws, columns: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(columns)
    for row in rows:
        ws.append([_cell_value(row.get(column, "")) for column in columns])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for idx, column in enumerate(columns, start=1):
        width = max(10, min(48, len(column) + 4))
        ws.column_dimensions[get_column_letter(idx)].width = width


def export_full_edif_xlsx(
    path: Path,
    tables: FullEdifTables,
    meta: dict[str, Any] | None = None,
    *,
    max_rows_per_sheet: int = 1_048_576,
) -> None:
    """Write full EDIF tables to an XLSX workbook.

    ``max_rows_per_sheet`` includes the header row. It is exposed for tests and
    for callers that want smaller chunks than Excel's hard limit.
    """

    if max_rows_per_sheet < 2:
        raise ValueError("max_rows_per_sheet must leave room for a header and data")

    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    data_rows_per_sheet = max_rows_per_sheet - 1

    for sheet_name in SHEET_ORDER:
        columns = FULL_EXPORT_COLUMNS[sheet_name]
        rows = tables.sheets.get(sheet_name, [])
        if sheet_name != "RawNodes" or len(rows) <= data_rows_per_sheet:
            ws = wb.create_sheet(sheet_name)
            _append_table(ws, columns, rows)
            continue

        for offset in range(0, len(rows), data_rows_per_sheet):
            chunk_index = offset // data_rows_per_sheet + 1
            ws = wb.create_sheet(_sheet_title(sheet_name, chunk_index))
            _append_table(ws, columns, rows[offset : offset + data_rows_per_sheet])

    if meta:
        ws = wb.create_sheet("_Meta")
        _append_table(
            ws,
            ["key", "value"],
            [{"key": key, "value": _cell_value(value)} for key, value in meta.items()],
        )

    wb.save(out)


def export_full_edif_json(
    path: Path,
    tables: FullEdifTables,
    meta: dict[str, Any] | None = None,
) -> None:
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    payload = tables.as_dict()
    if meta:
        payload["meta"] = meta
    out.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )
