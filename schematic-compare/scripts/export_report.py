from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from .diff_types import (
    ADD,
    COMPONENT_PROP,
    DELETE,
    NET_CONNECTION,
    NET_RENAME,
    PIN_INFO,
    PROPERTY_STYLE_TYPES,
    RENUMBER_REFDES,
)
from .models import DiffRow, merge_property_change_rows_for_display

_COMPARE_HEADERS: tuple[str, ...] = (
    "差异类型",
    "类别",
    "DSN 1",
    "DSN 2",
    "旧值",
    "新值",
    "页面",
    "修改详情",
    "风险等级",
    "差异说明",
)

_RISK_NOTE_MAP: dict[tuple[str, str], tuple[str, str]] = {
    ("器件", ADD): ("中", "新增器件，需核对功能"),
    ("器件", DELETE): ("中", "移除器件，需确认影响"),
    ("器件", RENUMBER_REFDES): ("低", "仅位号调整，属性不变"),
    ("器件", COMPONENT_PROP): ("中", "器件属性变更，请核对"),
    ("管脚", ADD): ("低", "新增管脚，请检查"),
    ("管脚", DELETE): ("低", "移除管脚，需确认"),
    ("管脚", RENUMBER_REFDES): ("低", "仅位号调整，无功能影响"),
    ("管脚", PIN_INFO): ("中", "管脚属性变更，请核对"),
    ("网络", ADD): ("中", "新增网络，需核对确认"),
    ("网络", DELETE): ("中", "移除网络，需核对确认"),
    ("网络", NET_RENAME): ("中", "网络重命名，请检查是否有掉网络"),
    ("网络", NET_CONNECTION): ("中", "连接变更，需核对确认"),
}

# 器件属性槽位 → (风险等级, 差异说明)，与产品判定表一致
_COMPONENT_SLOT_RISK_NOTE: dict[str, tuple[str, str]] = {
    "Value": ("中", "Value变更，影响电气特性"),
    "料号": ("中", "物料变更，需核对"),
    "描述": ("低", "器件描述变更，请检查"),
    "型号": ("中", "型号变更，需核对"),
    "品牌": ("低", "厂商变更，请检查"),
    "封装": ("高", "封装变更，需核对 PCB"),
    "贴装": ("高", "贴装变更，需核对 EBOM"),
    "电压": ("高", "耐压变更，存在电气风险"),
    "精度": ("高", "精度调整，影响电路性能"),
    "功率": ("高", "功率变更，存在失效风险"),
}

_SLOT_DISPLAY_LABELS: frozenset[str] = frozenset(_COMPONENT_SLOT_RISK_NOTE.keys())


def _risk_max(a: str, b: str) -> str:
    rank = {"低": 0, "中": 1, "高": 2}
    return a if rank.get(a, 1) >= rank.get(b, 1) else b


def _risk_max_list(levels: list[str]) -> str:
    if not levels:
        return "中"
    out = levels[0]
    for x in levels[1:]:
        out = _risk_max(out, x)
    return out


def _meta_full_path(raw: str) -> str:
    text = (raw or "").strip()
    if not text:
        return ""
    try:
        return str(Path(text).expanduser().resolve())
    except OSError:
        return text


def _meta_file_name(raw: str) -> str:
    text = (raw or "").strip()
    if not text:
        return ""
    try:
        return Path(text).name
    except (OSError, ValueError):
        return ""


def _append_dsn_meta(info: Any, old_path: str, new_path: str) -> None:
    info.append(("DSN 1 路径", _meta_full_path(old_path)))
    info.append(("DSN 1 文件名", _meta_file_name(old_path)))
    info.append(("DSN 2 路径", _meta_full_path(new_path)))
    info.append(("DSN 2 文件名", _meta_file_name(new_path)))


def _format_compare_table_cell(text: str | None) -> str:
    value = "" if text is None else str(text)
    if not value:
        return ""
    value = value.replace(" | ", " ; ")
    if value.strip() == "-":
        return " "
    parts = value.split(" ; ")
    out: list[str] = []
    for part in parts:
        stripped = part.strip()
        if stripped == "-":
            out.append(" ")
            continue
        if stripped.endswith("：-"):
            out.append(stripped[:-1] + " ")
            continue
        if stripped.endswith(":-"):
            out.append(stripped[:-1] + " ")
            continue
        out.append(part)
    return " ; ".join(out)


def _split_object_id(row: DiffRow) -> tuple[str, list[str]]:
    object_id = (row.object_id or "").strip()
    for prefix, kind in (("[Parts] ", "parts"), ("[Pins] ", "pins"), ("[Nets] ", "nets")):
        if object_id.startswith(prefix):
            rest = object_id[len(prefix) :].strip()
            return kind, [part.strip() for part in rest.split("|")]
    return "", [object_id]


def _dsn_pair(row: DiffRow) -> tuple[str, str]:
    kind, parts = _split_object_id(row)
    if row.change_type == RENUMBER_REFDES:
        meta = row.meta or {}
        return str(meta.get("oldRefdes") or "").strip() or "-", str(meta.get("newRefdes") or "").strip() or "-"
    if row.change_type == ADD:
        if kind == "pins":
            return "-", ".".join(parts[:2]).strip(".") or "-"
        return "-", parts[0] if parts and parts[0] else "-"
    if row.change_type == DELETE:
        if kind == "pins":
            return ".".join(parts[:2]).strip(".") or "-", "-"
        return parts[0] if parts and parts[0] else "-", "-"
    if row.change_type == NET_RENAME:
        meta = row.meta or {}
        if meta.get("csvNetRenamePinsPageValues"):
            o = str(meta.get("renameFlatNetOld") or "").strip()
            n = str(meta.get("renameFlatNetNew") or "").strip()
            return o or "-", n or "-"
        return str(row.old_value or "").strip() or "-", str(row.new_value or "").strip() or "-"
    if row.change_type == NET_CONNECTION:
        return parts[0] if parts and parts[0] else "-", parts[0] if parts and parts[0] else "-"
    if kind == "pins":
        disp = ".".join(parts[:2]).strip(".")
        return disp or "-", disp or "-"
    if parts and parts[0]:
        return parts[0], parts[0]
    return "-", "-"


def _page_display(row: DiffRow) -> str:
    meta = row.meta or {}
    left = "/".join(
        part
        for part in (
            str(meta.get("schematicNameDsn1") or "").strip(),
            str(meta.get("pageNameDsn1") or "").strip(),
        )
        if part
    )
    right = "/".join(
        part
        for part in (
            str(meta.get("schematicNameDsn2") or "").strip(),
            str(meta.get("pageNameDsn2") or "").strip(),
        )
        if part
    )
    if not left and not right:
        return ""
    return _format_compare_table_cell(f"{left} | {right}")


def _iter_part_prop_raw_headers(row: DiffRow) -> list[str]:
    """Extract property header names from a merged component property row."""
    pn = (row.prop_name or "").strip()
    if pn:
        return [pn]
    out: list[str] = []
    ov = str(row.old_value or "")
    if " | " in ov:
        for bit in ov.split(" | "):
            bit = bit.strip()
            if not bit:
                continue
            for sep in ("：", ":"):
                if sep in bit:
                    h = bit.split(sep, 1)[0].strip()
                    if h:
                        out.append(h)
                    break
        if out:
            return out
    det = str(row.detail or "")
    for seg in det.split("；"):
        seg = seg.strip()
        if not seg or "→" not in seg:
            continue
        left = seg.split("→", 1)[0].strip()
        sp = left.rfind(" ")
        if sp > 0:
            out.append(left[:sp].strip())
        elif left:
            out.append(left)
    return out


def _risk_level(row: DiffRow) -> str:
    r, _n = _auto_risk_and_diff_note(row)
    return r


def _diff_note(row: DiffRow) -> str:
    _r, n = _auto_risk_and_diff_note(row)
    return n


def _auto_risk_and_diff_note(row: DiffRow) -> tuple[str, str]:
    """Per-slot risk assessment matching source tool's _auto_risk_and_diff_note."""
    c = (row.category or "").strip()
    t = (row.change_type or "").strip()

    if c == "器件" and t == COMPONENT_PROP:
        risks: list[str] = []
        notes: list[str] = []
        seen_note: set[str] = set()
        for raw_h in _iter_part_prop_raw_headers(row):
            lab = (raw_h or "").strip()
            if lab in _COMPONENT_SLOT_RISK_NOTE:
                rk, msg = _COMPONENT_SLOT_RISK_NOTE[lab]
                risks.append(rk)
                if msg not in seen_note:
                    seen_note.add(msg)
                    notes.append(msg)
            else:
                risks.append("中")
                msg = "器件属性变更，请核对"
                if msg not in seen_note:
                    seen_note.add(msg)
                    notes.append(msg)
        if not risks:
            return "中", "器件属性变更，请核对"
        return _risk_max_list(risks), "；".join(notes)

    if c == "管脚" and t == PIN_INFO:
        pl = _pin_table_label_for_risk(row)
        has_pin = "脚名" in pl
        has_net = "网名" in pl
        rs: list[str] = []
        ns: list[str] = []
        if has_pin:
            rs.append("低")
            ns.append("仅名称调整，无功能影响")
        if has_net:
            rs.append("中")
            ns.append("请检查是否有掉网络")
        if rs:
            return _risk_max_list(rs), "；".join(ns)
        return "中", "管脚属性变更，请核对"

    return _RISK_NOTE_MAP.get((c, t), ("中", ""))


def _pin_table_label_for_risk(row: DiffRow) -> str:
    """Get pin property display labels for risk assessment."""
    pn = (row.prop_name or "").strip()
    if pn:
        return _value_label("管脚", pn)
    ov = str(row.old_value or "")
    if " | " in ov:
        labs: list[str] = []
        for bit in ov.split(" | "):
            bit = bit.strip()
            for sep in ("：", ":"):
                if sep in bit:
                    h = bit.split(sep, 1)[0].strip()
                    if h:
                        labs.append(_value_label("管脚", h))
                    break
        if labs:
            return "|".join(labs)
    return ""


def _value_label(category: str, prop_name: str | None) -> str:
    text = (prop_name or "").strip()
    cf = text.casefold()
    if category == "管脚":
        if "pin" in cf and "name" in cf:
            return "脚名"
        if "net" in cf or "signal" in cf:
            return "网名"
    if category == "网络" and "pin" in cf:
        return "Pins"
    return text


def _tokenize_pin_list(text: str) -> list[str]:
    return [part.strip() for part in str(text or "").split(",") if part.strip()]


def _fold_common_pins_to_com(old_value: str, new_value: str) -> tuple[str, str]:
    old_tokens = _tokenize_pin_list(old_value)
    new_tokens = _tokenize_pin_list(new_value)
    old_map = {token.casefold(): token for token in old_tokens}
    new_map = {token.casefold(): token for token in new_tokens}
    if max(len(old_map), len(new_map)) <= 6:
        return old_value, new_value
    common = set(old_map) & set(new_map)
    only_old = [old_map[key] for key in sorted(set(old_map) - common, key=str.lower)]
    only_new = [new_map[key] for key in sorted(set(new_map) - common, key=str.lower)]

    def render(tokens: list[str]) -> str:
        out = list(tokens)
        if common:
            out.append("COM")
        return ", ".join(out)

    return render(only_old), render(only_new)


def _old_new_values(row: DiffRow) -> tuple[str, str]:
    old_value = str(row.old_value or "")
    new_value = str(row.new_value or "")
    if row.change_type == NET_RENAME and row.category == "网络":
        meta = row.meta or {}
        if meta.get("csvNetRenamePinsPageValues"):
            pn = str(meta.get("renamePinsPagePropName") or "Pins (Page)").strip()
            label = _value_label(row.category, pn)
            old_text = _format_compare_table_cell(f"{label}：{old_value if old_value else ' '}")
            new_text = _format_compare_table_cell(f"{label}：{new_value if new_value else ' '}")
            return old_text or " ", new_text or " "
        label = _value_label(row.category, row.prop_name or "Pins (Page)")
        old_text = _format_compare_table_cell(f"{label}：{old_value if old_value else ' '}")
        new_text = _format_compare_table_cell(f"{label}：{new_value if new_value else ' '}")
        return old_text or " ", new_text or " "
    if row.change_type == NET_CONNECTION and row.category == "网络":
        label = _value_label(row.category, row.prop_name)
        old_value, new_value = _fold_common_pins_to_com(old_value, new_value)
        old_text = _format_compare_table_cell(f"{label}：{old_value if old_value else ' '}")
        new_text = _format_compare_table_cell(f"{label}：{new_value if new_value else ' '}")
        return old_text or " ", new_text or " "
    if row.category in ("管脚", "网络") and row.change_type in PROPERTY_STYLE_TYPES and row.prop_name:
        label = _value_label(row.category, row.prop_name)
        ov = old_value.strip()
        nv = new_value.strip()
        old_text = _format_compare_table_cell(f"{label}：{ov}" if ov and ov not in ("-", "—", "–") else " ")
        new_text = _format_compare_table_cell(f"{label}：{nv}" if nv and nv not in ("-", "—", "–") else " ")
        return old_text or " ", new_text or " "
    old_text = _format_compare_table_cell(old_value)
    new_text = _format_compare_table_cell(new_value)
    return old_text or " ", new_text or " "


def _normalize_detail_display(raw: str) -> str:
    """Normalize detail column: convert English column names to Chinese, apply display rules."""
    import re
    s = raw
    s = re.sub(r"(?i)Pins\s*\(\s*Page\s*\)", "Pins", s)
    s = s.replace("Pins（Page）", "Pins").replace("Pins（page）", "Pins")
    s = re.sub(r"(?i)Net\s+Name", "网名", s)
    s = re.sub(r"(?i)Pin\s+Name", "脚名", s)
    # Deduplicate rename description
    s = re.sub(
        r"(网络.+?→.+?，Pins不变)\s*[；：]\s*(?:Flat\s*Net|FlatNet|网名)\s+.+$",
        r"\1", s, count=1, flags=re.I | re.M,
    )
    return s


def _detail_display(row: DiffRow) -> str:
    raw = str(row.detail or "")
    if row.change_type == NET_CONNECTION and row.category == "网络":
        from .csv_diff import _nets_pins_connection_detail_cn

        old_value, new_value = _fold_common_pins_to_com(
            str(row.old_value or ""), str(row.new_value or "")
        )
        detail = _nets_pins_connection_detail_cn(old_value, new_value)
        return _normalize_detail_display(_format_compare_table_cell(detail))
    return _normalize_detail_display(_format_compare_table_cell(raw))


def export_markdown(
    path: str | Path,
    rows: list[DiffRow],
    meta: dict[str, Any] | None = None,
) -> None:
    output = Path(path)
    meta = meta or {}
    merged_rows = merge_property_change_rows_for_display(rows)

    lines = [
        "# 原理图对比报告",
        "",
        f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
    ]
    if meta.get("oldPath"):
        lines.append(f"DSN 1：`{meta['oldPath']}`")
    if meta.get("newPath"):
        lines.append(f"DSN 2：`{meta['newPath']}`")
    lines.append("")
    lines.append(f"差异总数：{len(merged_rows)}")
    lines.append("")
    lines.append("| 差异类型 | 类别 | DSN 1 | DSN 2 | 旧值 | 新值 | 页面 | 修改详情 | 风险等级 | 差异说明 |")
    lines.append("|---|---|---|---|---|---|---|---|---|---|")
    for row in merged_rows:
        dsn1, dsn2 = _dsn_pair(row)
        old_value, new_value = _old_new_values(row)
        lines.append(
            "| "
            + " | ".join(
                (
                    row.change_type,
                    row.category,
                    dsn1,
                    dsn2,
                    old_value,
                    new_value,
                    _page_display(row),
                    _detail_display(row),
                    _risk_level(row),
                    _diff_note(row),
                )
            )
            + " |"
        )
    output.write_text("\n".join(lines), encoding="utf-8")


def export_excel(
    path: str | Path,
    rows: list[DiffRow],
    meta: dict[str, Any] | None = None,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError("Excel 导出需要 openpyxl 库。请执行：pip install openpyxl") from exc

    output = Path(path)
    meta = meta or {}
    merged_rows = merge_property_change_rows_for_display(rows)

    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "对比结果"

    header_fill = PatternFill("solid", fgColor="D4D0C8")
    header_font = Font(bold=True)
    for col, header in enumerate(_COMPARE_HEADERS, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row_index, row in enumerate(merged_rows, start=2):
        dsn1, dsn2 = _dsn_pair(row)
        old_value, new_value = _old_new_values(row)
        values = (
            row.change_type,
            row.category,
            _format_compare_table_cell(dsn1),
            _format_compare_table_cell(dsn2),
            old_value,
            new_value,
            _page_display(row),
            _detail_display(row),
            _risk_level(row),
            _diff_note(row),
        )
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = (12, 10, 22, 22, 44, 44, 20, 40, 8, 22)
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

    info = workbook.create_sheet("元数据")
    info.append(("导出时间", datetime.now().isoformat(timespec="seconds")))
    _append_dsn_meta(info, str(meta.get("oldPath", "")), str(meta.get("newPath", "")))
    info.append(("对比属性", ", ".join(meta.get("selectedProps", []))))

    workbook.save(output)
