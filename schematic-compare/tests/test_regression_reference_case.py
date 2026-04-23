from __future__ import annotations

from collections import Counter
from pathlib import Path
import subprocess
import sys

import pytest
from openpyxl import load_workbook

from scripts.csv_diff import compare_all_dsn_csvs
from scripts.dsn_capture_export import _build_export_bundle_tcl
from scripts.export_report import export_excel
from scripts.models import DiffRow


FIXTURE_ROOT = Path("/Users/shenmingjie/tinno/hardware-diagram")
OLD_EDIF = FIXTURE_ROOT / "AI_SCH_CPU_V1.EDF"
NEW_EDIF = FIXTURE_ROOT / "AI_SCH_CPU_V2.EDF"


@pytest.mark.skipif(
    not OLD_EDIF.is_file() or not NEW_EDIF.is_file(),
    reason="reference EDIF files are not available on this machine",
)
def test_edif_reference_case_matches_expected_diff_distribution(tmp_path: Path) -> None:
    from scripts.edif_import import import_edif_pair_to_dirs

    old_dir = tmp_path / "v1"
    new_dir = tmp_path / "v2"
    old_dir.mkdir()
    new_dir.mkdir()

    out1, out2 = import_edif_pair_to_dirs(OLD_EDIF, NEW_EDIF, old_dir, new_dir)
    rows = compare_all_dsn_csvs(
        out1["parts"],
        out2["parts"],
        out1["pins"],
        out2["pins"],
        out1["nets"],
        out2["nets"],
    )

    counts = Counter((row.category, row.change_type) for row in rows)

    assert len(rows) == 77
    assert counts == Counter(
        {
            ("器件", "位号重编"): 54,
            ("器件", "删除"): 2,
            ("器件", "新增"): 3,
            ("管脚", "删除"): 6,
            ("管脚", "新增"): 6,
            ("网络", "新增"): 1,
            ("网络", "连接变更"): 5,
        }
    )


def test_export_excel_uses_standard_compare_columns(tmp_path: Path) -> None:
    rows = [
        DiffRow(
            category="器件",
            change_type="位号重编",
            object_id="[Parts] C9001|UTAH_MB|10_BB_PWR_PDN1",
            detail="位号C1001→C9001，属性不变",
            prop_name="refdes",
            old_value="Value：4.3uF | 料号：- | 描述：- | 型号：- | 品牌：- | 封装：- | 贴装：- | 电压：- | 精度：- | 功率：-",
            new_value="Value：4.3uF | 料号：- | 描述：- | 型号：- | 品牌：- | 封装：- | 贴装：- | 电压：- | 精度：- | 功率：-",
            meta={
                "oldRefdes": "C1001",
                "newRefdes": "C9001",
                "schematicNameDsn1": "UTAH_MB",
                "pageNameDsn1": "10_BB_PWR_PDN1",
                "schematicNameDsn2": "UTAH_MB",
                "pageNameDsn2": "10_BB_PWR_PDN1",
            },
        )
    ]
    out = tmp_path / "report.xlsx"

    export_excel(
        out,
        rows,
        {
            "oldPath": str(OLD_EDIF),
            "newPath": str(NEW_EDIF),
            "selectedProps": [
                "Description",
                "ISNC",
                "PCB Footprint",
                "Part Number",
                "Power",
                "Tolerance",
                "Value",
                "Vendor",
                "Vendor_PN",
                "Voltage",
            ],
        },
    )

    wb = load_workbook(out, read_only=True, data_only=True)
    ws = wb["对比结果"]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    first = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))

    assert header == (
        "差异类型",
        "类别",
        "DSN 1",
        "DSN 2",
        "旧值",
        "新值",
        "页面",
        "修改详情",
        "风险等级",
        "差异说明",
    )
    assert first == (
        "位号重编",
        "器件",
        "C1001",
        "C9001",
        "Value：4.3uF ; 料号：  ; 描述：  ; 型号：  ; 品牌：  ; 封装：  ; 贴装：  ; 电压：  ; 精度：  ; 功率： ",
        "Value：4.3uF ; 料号：  ; 描述：  ; 型号：  ; 品牌：  ; 封装：  ; 贴装：  ; 电压：  ; 精度：  ; 功率： ",
        "UTAH_MB/10_BB_PWR_PDN1 ; UTAH_MB/10_BB_PWR_PDN1",
        "位号C1001→C9001，属性不变",
        "低",
        "仅位号调整，属性不变",
    )


@pytest.mark.skipif(
    not OLD_EDIF.is_file() or not NEW_EDIF.is_file(),
    reason="reference EDIF files are not available on this machine",
)
def test_csv_cli_output_matches_source_export_for_same_csv_input(tmp_path: Path) -> None:
    source_root = Path(
        "/Users/shenmingjie/tinno/hardware-diagram/source-code/SchCompare_source_backup_20260423_112243"
    )
    if not source_root.is_dir():
        pytest.skip("reference source tree is not available on this machine")

    from scripts.edif_import import import_edif_pair_to_dirs

    old_dir = tmp_path / "v1"
    new_dir = tmp_path / "v2"
    old_dir.mkdir()
    new_dir.mkdir()
    out1, out2 = import_edif_pair_to_dirs(OLD_EDIF, NEW_EDIF, old_dir, new_dir)

    cli_report = tmp_path / "cli_csv.xlsx"
    subprocess.run(
        [
            sys.executable,
            "-m",
            "scripts.schcompare_cli",
            "csv",
            str(old_dir),
            str(new_dir),
            "-o",
            str(cli_report),
        ],
        check=True,
        cwd="/Users/shenmingjie/.codex/skills/schematic-compare",
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )

    sys.path.insert(0, str(source_root))
    from app.csv_diff import compare_all_dsn_csvs as source_compare  # type: ignore
    from app.export_excel import export_rows as source_export  # type: ignore

    source_report = tmp_path / "source_csv.xlsx"
    source_rows, err = source_compare(
        {
            "lastDsn1PartsCsv": str(out1["parts"]),
            "lastDsn2PartsCsv": str(out2["parts"]),
            "lastDsn1PinsCsv": str(out1["pins"]),
            "lastDsn2PinsCsv": str(out2["pins"]),
            "lastDsn1NetsCsv": str(out1["nets"]),
            "lastDsn2NetsCsv": str(out2["nets"]),
            "oldPath": str(OLD_EDIF),
            "newPath": str(NEW_EDIF),
            "pinsCsvColumnCompare": [
                {"on": True, "d1": "Pin Name", "d2": "Pin Name"},
                {"on": False, "d1": "Net Name", "d2": "Net Name"},
            ],
            "netsCsvColumnCompare": [
                {"on": True, "d1": "Pins (Page)", "d2": "Pins (Page)"}
            ],
        }
    )
    assert err is None
    source_export(
        source_report,
        source_rows,
        {
            "oldPath": str(OLD_EDIF),
            "newPath": str(NEW_EDIF),
            "selectedProps": [
                "Description",
                "ISNC",
                "PCB Footprint",
                "Part Number",
                "Power",
                "Tolerance",
                "Value",
                "Vendor",
                "Vendor_PN",
                "Voltage",
            ],
        },
    )

    def read_compare_rows(path: Path) -> list[tuple[str, ...]]:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb["对比结果"]
        return [
            tuple("" if cell is None else str(cell) for cell in row)
            for row in ws.iter_rows(values_only=True)
        ]

    assert read_compare_rows(cli_report) == read_compare_rows(source_report)


def test_dsn_bundle_script_sources_tcl_once_and_exports_both_slots(tmp_path: Path) -> None:
    dsn1 = tmp_path / "old.dsn"
    dsn2 = tmp_path / "new.dsn"
    out1 = tmp_path / "export1"
    out2 = tmp_path / "export2"
    script = _build_export_bundle_tcl([(dsn1, out1), (dsn2, out2)])

    assert script.count("source ") == 2
    assert script.count("exportDsnCsv ") == 2
    assert str(dsn1.resolve().as_posix()) in script
    assert str(dsn2.resolve().as_posix()) in script
    assert str(out1.resolve().as_posix()) in script
    assert str(out2.resolve().as_posix()) in script


def test_dsn_tcl_files_match_reference_source_when_available() -> None:
    source_root = Path(
        "/Users/shenmingjie/tinno/hardware-diagram/source-code/SchCompare_source_backup_20260423_112243"
    )
    if not source_root.is_dir():
        pytest.skip("reference source tree is not available on this machine")

    current_export = Path(
        "/Users/shenmingjie/.codex/skills/schematic-compare/scripts/tcl/capExportAllInfo.tcl"
    )
    current_find = Path(
        "/Users/shenmingjie/.codex/skills/schematic-compare/scripts/tcl/capFind.tcl"
    )
    source_export = source_root / "capExportAllInfo.tcl"
    source_find = source_root / "capFind.tcl"

    assert current_export.read_text(encoding="utf-8") == source_export.read_text(
        encoding="utf-8"
    )
    assert current_find.read_text(encoding="utf-8") == source_find.read_text(
        encoding="utf-8"
    )
